# Copyright (c) 2024, Resilient Tech and Contributors
# See license.txt

from unittest.mock import patch

import openpyxl
from frappe.tests import IntegrationTestCase, UnitTestCase

from india_compliance.gst_india.doctype.gstr_1.gstr_1_export import (
    GovExcel,
    _filter_data_by_section,
    _get_excel_sheet_names,
    _get_gov_filename,
    _get_hsn_sections,
    _get_selected_sections,
)
from india_compliance.gst_india.utils.exporter import ExcelExporter
from india_compliance.gst_india.utils.gstr_1 import (
    JSON_CATEGORY_EXCEL_CATEGORY_MAPPING,
    GovExcelSheetName,
    GovJsonKey,
)

# Sections the per-section Excel endpoint accepts: every GovJsonKey value that maps
# to a sheet in JSON_CATEGORY_EXCEL_CATEGORY_MAPPING (e.g. supeco/sec_sum are excluded).
GOV_EXCEL_SECTIONS = frozenset(
    key.value for key in GovJsonKey if key.value in JSON_CATEGORY_EXCEL_CATEGORY_MAPPING
)


class TestGSTR1(IntegrationTestCase):
    pass


class TestFilterDataBySection(UnitTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.gov_data = {
            "b2b": [{"invoice": "INV-001"}],
            "cdnr": [{"note": "CN-001"}],
            "b2cs": [{"supply": "S-001"}],
        }

    def test_returns_all_sections_when_section_is_none(self):
        result = _filter_data_by_section(self.gov_data, None)
        self.assertEqual(result, self.gov_data)

    def test_returns_matching_section(self):
        result = _filter_data_by_section(self.gov_data, "b2b")
        self.assertEqual(result, {"b2b": [{"invoice": "INV-001"}]})
        self.assertNotIn("cdnr", result)
        self.assertNotIn("b2cs", result)

    def test_returns_empty_for_unknown_section(self):
        result = _filter_data_by_section(self.gov_data, "nonexistent")
        self.assertEqual(result, {})


class TestSectionDataKeys(UnitTestCase):
    def test_non_hsn_section_returns_single_key(self):
        self.assertEqual(_get_selected_sections("b2b", is_hsn_bifurcated=False), ("b2b",))

    def test_hsn_pre_bifurcation_returns_single_hsn_key(self):
        self.assertEqual(_get_selected_sections(GovJsonKey.HSN.value, is_hsn_bifurcated=False), ("hsn",))

    def test_hsn_post_bifurcation_returns_split_keys(self):
        self.assertEqual(
            _get_selected_sections(GovJsonKey.HSN.value, is_hsn_bifurcated=True),
            ("hsn_b2b", "hsn_b2c"),
        )

    def test_unknown_section_is_returned_as_is(self):
        self.assertEqual(_get_selected_sections("nonexistent", is_hsn_bifurcated=False), ("nonexistent",))


class TestSectionSheetNames(UnitTestCase):
    def test_non_hsn_section_returns_single_sheet(self):
        self.assertEqual(
            _get_excel_sheet_names("b2b", is_hsn_bifurcated=False),
            (GovExcelSheetName.B2B.value,),
        )

    def test_hsn_pre_bifurcation_returns_single_sheet(self):
        self.assertEqual(
            _get_excel_sheet_names(GovJsonKey.HSN.value, is_hsn_bifurcated=False),
            (GovExcelSheetName.HSN.value,),
        )

    def test_hsn_post_bifurcation_returns_both_split_sheets(self):
        self.assertEqual(
            _get_excel_sheet_names(GovJsonKey.HSN.value, is_hsn_bifurcated=True),
            (GovExcelSheetName.HSN_B2B.value, GovExcelSheetName.HSN_B2C.value),
        )

    def test_supeco_resolves_to_eco_sheet(self):
        self.assertEqual(
            _get_excel_sheet_names(GovJsonKey.SUPECOM.value, is_hsn_bifurcated=False),
            (GovExcelSheetName.ECO.value,),
        )

    def test_unknown_section_returns_empty_tuple(self):
        self.assertEqual(_get_excel_sheet_names("nonexistent", is_hsn_bifurcated=False), ())


class TestGetGovFilename(UnitTestCase):
    GSTIN = "29AABCU9603R1ZM"
    PERIOD = "032024"

    def test_includes_section_name_when_section_given(self):
        filename = _get_gov_filename(self.GSTIN, self.PERIOD, "b2b")
        self.assertEqual(filename, f"GSTR-1-Gov-{self.GSTIN}-{self.PERIOD}-b2b")

    def test_default_filename_when_no_section(self):
        filename = _get_gov_filename(self.GSTIN, self.PERIOD, None)
        self.assertEqual(filename, f"GSTR-1-Gov-{self.GSTIN}-{self.PERIOD}")


class TestGovExcelSectionSheets(UnitTestCase):
    """
    Static consistency check: every sheet name we promise to keep under a section
    filter must exist in the Gov Excel template for that version. Catches drift
    if a future GSTN template renames or removes a sheet.
    """

    def _assert_sections_against_template(self, template_path, is_hsn_bifurcated):
        wb = openpyxl.load_workbook(template_path, read_only=True)
        template_sheets = set(wb.sheetnames)
        wb.close()

        for section in GOV_EXCEL_SECTIONS:
            expected = set(_get_excel_sheet_names(section, is_hsn_bifurcated))
            missing = expected - template_sheets
            self.assertFalse(
                missing,
                f"Section {section!r} (bifurcated={is_hsn_bifurcated}) references "
                f"sheets missing from template {template_path}: {missing}",
            )

        self.assertIn(
            GovExcelSheetName.MASTER.value,
            template_sheets,
            f"Template {template_path} is missing the 'master' reference sheet",
        )

    def test_pre_bifurcation_template(self):
        self._assert_sections_against_template(GovExcel.TEMPLATE_EXCEL_FILE["V2.0"], is_hsn_bifurcated=False)

    def test_post_bifurcation_template(self):
        self._assert_sections_against_template(GovExcel.TEMPLATE_EXCEL_FILE["V2.1"], is_hsn_bifurcated=True)


class TestGovExcelBuildSection(UnitTestCase):
    """
    End-to-end: drive GovExcel.build_excel against the real template files
    and assert the resulting workbook contains exactly the sheets we promised.

    This catches the three regressions that the static checks above missed:
      - HSN dead-code branch (wrong enum compared)
      - string-iteration on `sections=section`
      - missing `supeco → eco` mapping
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.gov = GovExcel()
        cls.gov.gstin = "29AABCU9603R1ZM"
        cls.gov.period = "032024"

    def _build(self, template_version, section, is_hsn_bifurcated):
        captured = {}

        def fake_export(exporter_self, file_name):
            captured["sheetnames"] = list(exporter_self.wb.sheetnames)
            captured["file_name"] = file_name

        self.gov.is_hsn_bifurcated = is_hsn_bifurcated
        with patch.object(ExcelExporter, "export", fake_export):
            self.gov.build_excel(
                data={},
                file=GovExcel.TEMPLATE_EXCEL_FILE[template_version],
                section=section,
            )
        return captured

    def test_v20_b2b_keeps_only_b2b_and_master(self):
        captured = self._build("V2.0", "b2b", is_hsn_bifurcated=False)
        self.assertEqual(
            set(captured["sheetnames"]),
            {GovExcelSheetName.MASTER.value, GovExcelSheetName.B2B.value},
        )

    def test_v20_hsn_keeps_single_hsn_sheet(self):
        captured = self._build("V2.0", "hsn", is_hsn_bifurcated=False)
        self.assertEqual(
            set(captured["sheetnames"]),
            {GovExcelSheetName.MASTER.value, GovExcelSheetName.HSN.value},
        )

    def test_v21_hsn_keeps_both_bifurcated_sheets(self):
        captured = self._build("V2.1", "hsn", is_hsn_bifurcated=True)
        self.assertEqual(
            set(captured["sheetnames"]),
            {
                GovExcelSheetName.MASTER.value,
                GovExcelSheetName.HSN_B2B.value,
                GovExcelSheetName.HSN_B2C.value,
            },
        )

    def test_v21_supeco_keeps_eco_sheet(self):
        captured = self._build("V2.1", "supeco", is_hsn_bifurcated=True)
        self.assertEqual(
            set(captured["sheetnames"]),
            {GovExcelSheetName.MASTER.value, GovExcelSheetName.ECO.value},
        )

    def test_filename_includes_section_suffix(self):
        captured = self._build("V2.1", "cdnr", is_hsn_bifurcated=True)
        self.assertTrue(
            captured["file_name"].endswith("-cdnr"),
            f"Expected filename to end with '-cdnr', got {captured['file_name']!r}",
        )

    def test_no_section_keeps_all_template_sheets(self):
        captured = self._build("V2.1", section=None, is_hsn_bifurcated=True)
        # Sanity: at minimum master + a few core sheets remain when no filter applied.
        self.assertIn(GovExcelSheetName.MASTER.value, captured["sheetnames"])
        self.assertIn(GovExcelSheetName.B2B.value, captured["sheetnames"])
        self.assertIn(GovExcelSheetName.ECO.value, captured["sheetnames"])
        self.assertFalse(captured["file_name"].endswith("-None"))
